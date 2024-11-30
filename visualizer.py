import pandas as pd
import matplotlib.pyplot as plt
from mplfonts import use_font
import io
import base64
from datetime import datetime
import os
import json
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Border, Side, Alignment

use_font('Noto Serif CJK SC')

class DataVisualizer:
    def __init__(self, json_file='lazada_products.json', output_dir='analysis_results'):
        self.json_file = json_file
        self.output_dir = output_dir
        self.data = None
        self.stats = {}

        import os
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)

    def load_data(self):
        try:
            self.data = pd.read_json(self.json_file)
            self._clean_data()
            self.create_visualizations()
            self.calculate_statistics()
            self.save_analysis()
            print(f"Successfully analyzed {len(self.data)} records")
        except Exception as e:
            print(f"Error: {str(e)}")

    def _clean_data(self):
        def clean_price(price):
            try:
                return float(price.replace('$', '').replace(',', ''))
            except:
                return None

        def clean_rating(rating):
            try:
                if rating == 'N/A':
                    return None
                return float(rating.split()[0])
            except:
                return None

        def clean_sold(sold):
            try:
                if sold == 'N/A':
                    return 0
                sold = sold.lower()
                if 'k' in sold:
                    return int(float(sold.replace('k', '').replace(' sold', '')) * 1000)
                return int(''.join(filter(str.isdigit, sold)))
            except:
                return 0

        self.data['price_clean'] = self.data['price'].apply(clean_price)
        self.data['rating_clean'] = self.data['rating'].apply(clean_rating)
        self.data['sold_clean'] = self.data['sold'].apply(clean_sold)

    def create_visualizations(self):
        plt.style.use('default')  

        fig1, (ax1, ax2) = plt.subplots(1, 2, figsize=(15, 6))

        ax1.hist(self.data['price_clean'].dropna(), bins=30, edgecolor='black', alpha=0.7)
        ax1.set_title('Price Distribution')
        ax1.set_xlabel('Price ($)')
        ax1.set_ylabel('Count')

        ax2.boxplot(self.data['price_clean'].dropna())
        ax2.set_title('Price Boxplot')
        ax2.set_ylabel('Price ($)')
        
        plt.tight_layout()
        plt.savefig(f'{self.output_dir}/price_analysis.png')
        plt.close()

        plt.figure(figsize=(10, 6))
        plt.scatter(self.data['price_clean'], self.data['rating_clean'], alpha=0.6)
        plt.title('Price vs Rating')
        plt.xlabel('Price ($)')
        plt.ylabel('Rating')
        plt.savefig(f'{self.output_dir}/price_vs_rating.png')
        plt.close()

        plt.figure(figsize=(12, 6))
        location_counts = self.data['location'].value_counts()
        plt.barh(range(len(location_counts)), location_counts.values)
        plt.yticks(range(len(location_counts)), location_counts.index)
        plt.title('Products by Location')
        plt.xlabel('Number of Products')
        plt.tight_layout()
        plt.savefig(f'{self.output_dir}/location_distribution.png')
        plt.close()

        plt.figure(figsize=(10, 6))
        plt.scatter(self.data['price_clean'], self.data['sold_clean'], alpha=0.6)
        plt.title('Price vs Sales')
        plt.xlabel('Price ($)')
        plt.ylabel('Units Sold')
        plt.savefig(f'{self.output_dir}/price_vs_sales.png')
        plt.close()

    def calculate_statistics(self):
        price_stats = self.data['price_clean'].describe()
        rating_stats = self.data['rating_clean'].describe()
        sales_stats = self.data['sold_clean'].describe()
        correlations = self.data[['price_clean', 'rating_clean', 'sold_clean']].corr()
        top_selling = self.data.nlargest(5, 'sold_clean')[['name', 'price', 'sold', 'rating']]
        location_dist = self.data['location'].value_counts()

        self.stats = {
            'analysis_date': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'total_products': len(self.data),
            'price_statistics': price_stats.to_dict(),
            'rating_statistics': rating_stats.to_dict(),
            'sales_statistics': sales_stats.to_dict(),
            'correlations': correlations.to_dict(),
            'top_selling_products': top_selling.to_dict('records'),
            'location_distribution': location_dist.to_dict()
        }

    def save_analysis(self):
        def format_value(value):
            """Helper function to format values properly"""
            if isinstance(value, str):
                return value
            if pd.isna(value) or value == "NaN":
                return "N.A"
            if isinstance(value, (int, float)):
                return round(value, 2)
            return value

        def clean_stats_dict(d):
            """Recursively clean dictionary of NaN values"""
            if isinstance(d, dict):
                return {k: clean_stats_dict(v) for k, v in d.items()}
            elif isinstance(d, list):
                return [clean_stats_dict(v) for v in d]
            else:
                return format_value(d)

        stats_copy = clean_stats_dict(self.stats)

        with open(f'{self.output_dir}/statistics.json', 'w') as f:
            json.dump(stats_copy, f, indent=4)

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Analysis Report"

            # Styling
            title_font = Font(bold=True, size=14)
            header_font = Font(bold=True)
            border = Border(bottom=Side(style='thin'))

            # Title Section
            ws['A1'] = 'Lazada Product Analysis Report'
            ws['A1'].font = title_font
            ws['A2'] = 'Generated on:'
            ws['B2'] = stats_copy['analysis_date']
            ws.merge_cells('A1:D1')

            current_row = 4

            # Product Overview Section
            ws[f'A{current_row}'] = 'Product Overview'
            ws[f'A{current_row}'].font = header_font
            current_row += 1
            ws[f'A{current_row}'] = 'Total Products'
            ws[f'B{current_row}'] = stats_copy['total_products']
            current_row += 2

            # Price Statistics Section
            ws[f'A{current_row}'] = 'Price Statistics (SGD)'
            ws[f'A{current_row}'].font = header_font
            current_row += 1
            stats_headers = ['Metric', 'Value']
            for col, header in enumerate(stats_headers, 1):
                ws.cell(row=current_row, column=col, value=header).font = header_font
            current_row += 1
            
            price_stats = stats_copy['price_statistics']
            metrics = [
                ('Average', price_stats['mean']),
                ('Median', price_stats['50%']),
                ('Minimum', price_stats['min']),
                ('Maximum', price_stats['max']),
                ('Standard Deviation', price_stats['std'])
            ]
            for metric, value in metrics:
                ws[f'A{current_row}'] = metric
                ws[f'B{current_row}'] = value
                current_row += 1
            current_row += 1

            # Sales Statistics Section
            ws[f'A{current_row}'] = 'Sales Statistics'
            ws[f'A{current_row}'].font = header_font
            current_row += 1
            sales_stats = stats_copy['sales_statistics']
            metrics = [
                ('Average Sales', sales_stats['mean']),
                ('Median Sales', sales_stats['50%']),
                ('Minimum Sales', sales_stats['min']),
                ('Maximum Sales', sales_stats['max']),
                ('Total Products Sold', sales_stats['count'])
            ]
            for metric, value in metrics:
                ws[f'A{current_row}'] = metric
                ws[f'B{current_row}'] = value
                current_row += 1
            current_row += 1

            # Top Selling Products Section
            ws[f'A{current_row}'] = 'Top Selling Products'
            ws[f'A{current_row}'].font = header_font
            current_row += 1
            headers = ['Product Name', 'Price (SGD)', 'Units Sold', 'Rating']
            for col, header in enumerate(headers, 1):
                ws.cell(row=current_row, column=col, value=header).font = header_font
            current_row += 1
            
            for product in stats_copy['top_selling_products']:
                ws.append([
                    product['name'],
                    product['price'],
                    product['sold'],
                    product['rating']
                ])
            current_row += len(stats_copy['top_selling_products']) + 1

            # Correlation Analysis Section
            ws[f'A{current_row}'] = 'Correlation Analysis'
            ws[f'A{current_row}'].font = header_font
            current_row += 1
            headers = ['', 'Price', 'Rating', 'Sales']
            for col, header in enumerate(headers, 1):
                ws.cell(row=current_row, column=col, value=header).font = header_font
            current_row += 1
            
            correlations = stats_copy['correlations']
            metrics = ['price_clean', 'rating_clean', 'sold_clean']
            for metric in metrics:
                row = [metric.replace('_clean', '').capitalize()]
                for col_metric in metrics:
                    value = correlations[metric][col_metric]
                    row.append(value)
                ws.append(row)
            current_row += 5

            # Add visualizations
            charts_section = current_row
            image_files = [
                ('Price vs Sales Analysis', 'price_vs_sales.png', f'A{charts_section}'),
                ('Price Distribution', 'price_analysis.png', f'E{charts_section}'),
                ('Location Distribution', 'location_distribution.png', f'A{charts_section + 20}'),
                ('Price vs Rating', 'price_vs_rating.png', f'E{charts_section + 20}')
            ]

            for title, filename, cell in image_files:
                try:
                    img_path = f'{self.output_dir}/{filename}'
                    if os.path.exists(img_path):
                        ws[cell] = title
                        ws[cell].font = header_font
                        img = Image(img_path)
                        img.width = 400
                        img.height = 300
                        ws.add_image(img, cell)
                except Exception as e:
                    print(f"Could not add image {filename}: {str(e)}")

            default_widths = {
                'A': 40,  # Product names and metrics
                'B': 15,  # Values
                'C': 15,  # Additional data
                'D': 15,  # Additional data
                'E': 40,  # For second column of charts
            }
            
            for column, width in default_widths.items():
                ws.column_dimensions[column].width = width

            excel_path = f'{self.output_dir}/analysis_report.xlsx'
            wb.save(excel_path)
            print(f"Excel report saved to: {excel_path}")

        except Exception as e:
            print(f"Error creating Excel report: {str(e)}")

if __name__ == "__main__":
    visualizer = DataVisualizer()
    visualizer.load_data()
